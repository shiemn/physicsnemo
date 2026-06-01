#!/usr/bin/env python3
"""Export WandB CorrDiff eval runs to a color-coded Excel spreadsheet.

Usage:
    # First export — filter by date, save to file
    python wandb_to_excel.py --after 2026-05-10 --output runs.xlsx

    # Update an existing file — fetches only new runs since last export
    python wandb_to_excel.py --update runs.xlsx

    # Show delta columns vs. a named baseline run
    python wandb_to_excel.py --after 2026-05-01 --baseline my_baseline_run_tag

    # Select extra metrics and filter by run tag pattern
    python wandb_to_excel.py --update runs.xlsx --metrics rmse,crps,mppe,hrre_10.0mm
    python wandb_to_excel.py --after 2026-05-01 --tag-filter "norway_v2"
"""

import argparse
import datetime
import json
import os
import re
import sys

import wandb
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ALL_METRICS = [
    "rmse", "crps", "bias", "mae", "pcc",
    "spread", "skill", "spread_skill_ratio",
    "rmse_95th", "crps_95th", "rmse_w95th", "crps_w95th",
    "twcrps_5.0mm", "twcrps_10.0mm", "hrre_10.0mm", "mppe", "n_samples",
]
DEFAULT_METRICS = ["rmse", "crps", "bias", "pcc", "spread_skill_ratio"]

# Directional metrics — drives color-scale orientation and rank computation.
# bias: lower signed bias preferred for precipitation (underestimation < overestimation)
LOWER_BETTER = {
    "rmse", "crps", "bias", "mae",
    "rmse_95th", "crps_95th", "rmse_w95th", "crps_w95th",
    "twcrps_5.0mm", "twcrps_10.0mm", "hrre_10.0mm", "mppe",
}
HIGHER_BETTER = {"pcc", "spread_skill_ratio"}
# spread, skill, n_samples — neutral, no gradient applied

META_KEYS = [
    "run_name", "run_tag", "date",
    "guidance_scale", "num_ensembles", "sampler_type", "sampler_steps",
    "reg_ckpt", "res_ckpt",
]

GREEN        = "FF63BE7B"
WHITE        = "FFFFFFFF"
RED          = "FFF8696B"
HEADER_FILL  = PatternFill("solid", fgColor="FF4472C4")
HEADER_FONT  = Font(bold=True, color="FFFFFFFF")
FOOTER_FILL  = PatternFill("solid", fgColor="FFD9E1F2")
NEW_META_FILL = PatternFill("solid", fgColor="FFFFE599")  # yellow for new rows (metadata cols)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(
        description="Export WandB CorrDiff eval runs to a color-coded Excel spreadsheet.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"Available metrics: {', '.join(ALL_METRICS)}",
    )
    p.add_argument("--after",  type=datetime.date.fromisoformat, metavar="DATE",
                   help="Include runs created after DATE (YYYY-MM-DD)")
    p.add_argument("--before", type=datetime.date.fromisoformat, metavar="DATE",
                   help="Include runs created before DATE (YYYY-MM-DD)")
    p.add_argument("--tag-filter", metavar="PATTERN",
                   help="Regex filter on run_tag (client-side)")
    p.add_argument("--state", default="finished",
                   choices=["finished", "running", "crashed", "all"],
                   help="WandB run state filter (default: finished)")
    p.add_argument("--project", default="evaluation",
                   help="WandB project name (default: evaluation)")
    p.add_argument("--entity", default="shiemn",
                   help="WandB entity (default: shiemn)")
    p.add_argument("--metrics", default=",".join(DEFAULT_METRICS),
                   help=f"Comma-separated metrics to include "
                        f"(default: {','.join(DEFAULT_METRICS)})")
    p.add_argument("--output", default=None, metavar="FILE",
                   help="Output .xlsx path (default: wandb_eval_runs_YYYY-MM-DD.xlsx)")
    p.add_argument("--max-runs", type=int, default=200,
                   help="Max runs to fetch newest-first (default: 200)")
    p.add_argument("--baseline", metavar="TAG",
                   help="run_tag of baseline run; appends Δ delta columns for all others")
    p.add_argument("--update", metavar="FILE",
                   help="Update existing .xlsx in place: append only new runs since last export")
    return p.parse_args()


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------

def get_nested(d, *keys, default=None):
    """Traverse nested dict, falling back to slash-joined flat key lookup."""
    cur = d
    for k in keys:
        if not isinstance(cur, dict):
            cur = None
            break
        cur = cur.get(k)
    if cur is not None:
        return cur
    return d.get("/".join(keys), default)


def safe_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        return None if f != f else f  # drop NaN
    except (TypeError, ValueError):
        return None


def _metric_direction(m_name):
    """Returns True (lower=better), False (higher=better), or None (neutral)."""
    if m_name in LOWER_BETTER:
        return True
    if m_name in HIGHER_BETTER:
        return False
    return None


# ---------------------------------------------------------------------------
# WandB fetching
# ---------------------------------------------------------------------------

def fetch_runs(entity, project, state, after, before, max_runs):
    api = wandb.Api()
    filters = {"group": "CorrDiff-Eval"}
    if state != "all":
        filters["state"] = state
    if after:
        filters["createdAt"] = {"$gte": after.isoformat() + "T00:00:00"}
    if before:
        filters.setdefault("createdAt", {})["$lte"] = before.isoformat() + "T23:59:59"

    print(f"  Querying {entity}/{project} — state={state}, filters={filters}", flush=True)
    runs_iter = api.runs(
        path=f"{entity}/{project}",
        filters=filters,
        order="-created_at",
        per_page=min(max_runs, 200),
    )
    runs = []
    for run in runs_iter:
        runs.append(run)
        if len(runs) >= max_runs:
            break
    print(f"  Got {len(runs)} run(s) from WandB.", flush=True)
    return runs


def extract_row(run, selected_metrics):
    """Extract a flat row dict from a WandB Run object."""
    cfg     = run.config or {}
    summary = dict(run.summary) if run.summary else {}

    row = {
        "run_name":      run.name or "",
        "run_tag":       cfg.get("run_tag", ""),
        "date":          (run.created_at or "")[:10],
        "_url":          run.url or "",
        "guidance_scale": safe_float(get_nested(cfg, "generation", "guidance_scale")),
        "num_ensembles":  get_nested(cfg, "generation", "num_ensembles"),
        "sampler_type":   str(get_nested(cfg, "sampler", "type") or ""),
        "sampler_steps":  get_nested(cfg, "sampler", "num_steps"),
        "reg_ckpt": os.path.basename(str(
            get_nested(cfg, "generation", "io", "reg_ckpt_filename") or "")),
        "res_ckpt": os.path.basename(str(
            get_nested(cfg, "generation", "io", "res_ckpt_filename") or "")),
    }

    for prefix in ("regression", "diffusion"):
        for m in selected_metrics:
            row[f"{prefix}/{m}"] = safe_float(summary.get(f"{prefix}/{m}"))

    return row


# ---------------------------------------------------------------------------
# Excel reading (for --update)
# ---------------------------------------------------------------------------

def load_existing(filepath):
    """Read back rows + metadata from an xlsx previously written by this script.

    Returns (rows_list, last_updated_iso_str, saved_metrics_list).
    Rows are flat dicts with keys matching col_headers from write_excel.
    """
    wb = load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    run_name_col_idx = headers.index("run_name") if "run_name" in headers else 0

    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=False):
        # Blank cell in column 1 = separator before footer
        if not row_cells[0].value:
            break
        row_dict = {}
        for header, cell in zip(headers, row_cells):
            if header is None:
                continue
            val = cell.value
            # Skip formula strings (footer rows shouldn't appear here, but be safe)
            if isinstance(val, str) and val.startswith("="):
                val = None
            row_dict[header] = val

        # Recover WandB URL from the hyperlink set on run_name cell
        rn_cell = row_cells[run_name_col_idx]
        row_dict["_url"] = (rn_cell.hyperlink.target if rn_cell.hyperlink else "") or ""
        rows.append(row_dict)

    last_updated = None
    saved_metrics = None
    if "_meta" in wb.sheetnames:
        ms = wb["_meta"]
        last_updated = ms["A1"].value
        raw = ms["A2"].value
        if raw:
            try:
                saved_metrics = json.loads(raw)
            except Exception:
                pass

    return rows, last_updated, saved_metrics


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

def make_color_rule(lower_better):
    """3-color scale: white at median, green=good, red=bad."""
    start_color = GREEN if lower_better else RED
    end_color   = RED   if lower_better else GREEN
    return ColorScaleRule(
        start_type="min",      start_color=start_color,
        mid_type="percentile", mid_value=50, mid_color=WHITE,
        end_type="max",        end_color=end_color,
    )


def compute_ranks(rows, selected_metrics):
    """Return {prefix: {run_name: rank}} ranked by the primary metric.

    Primary metric = first LOWER_BETTER metric in selected_metrics,
    or first HIGHER_BETTER metric if none are lower-better.
    """
    result = {}
    for prefix in ("regression", "diffusion"):
        primary = next(
            (m for m in selected_metrics if m in LOWER_BETTER), None
        ) or next(
            (m for m in selected_metrics if m in HIGHER_BETTER), None
        )
        if primary is None:
            result[prefix] = {}
            continue
        lower_b = primary in LOWER_BETTER
        pairs = [
            (r["run_name"], r.get(f"{prefix}/{primary}"))
            for r in rows
            if r.get(f"{prefix}/{primary}") is not None
        ]
        pairs.sort(key=lambda x: x[1], reverse=not lower_b)
        result[prefix] = {name: i + 1 for i, (name, _) in enumerate(pairs)}
    return result


def write_excel(rows, selected_metrics, baseline_tag, output_path, new_run_names=None):
    if new_run_names is None:
        new_run_names = set()

    # Find baseline row for delta columns
    baseline_row = None
    if baseline_tag:
        baseline_row = next((r for r in rows if r.get("run_tag") == baseline_tag), None)
        if baseline_row is None:
            print(f"  Warning: baseline run_tag '{baseline_tag}' not found — skipping Δ columns.")

    # Build column layout:
    # [META_KEYS] [reg_rank] [regression/m...] [diff_rank] [diffusion/m...] [Δreg/m... Δdiff/m...]
    col_keys    = list(META_KEYS)   # dict-key for row lookups
    col_headers = list(META_KEYS)   # displayed header text

    for prefix, label in [("regression", "reg"), ("diffusion", "diff")]:
        col_keys.append(f"_rank_{prefix}")
        col_headers.append(f"{label}_rank")
        for m in selected_metrics:
            col_keys.append(f"{prefix}/{m}")
            col_headers.append(f"{prefix}/{m}")

    if baseline_row:
        for prefix in ("regression", "diffusion"):
            for m in selected_metrics:
                col_keys.append(f"_delta_{prefix}/{m}")
                col_headers.append(f"Δ{prefix}/{m}")

    n_meta = len(META_KEYS)

    ranks = compute_ranks(rows, selected_metrics)

    wb = Workbook()
    ws = wb.active
    ws.title = "Eval Runs"

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, header in enumerate(col_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # ── Data rows ────────────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=2):
        is_new = row_data.get("run_name", "") in new_run_names

        for col_idx, key in enumerate(col_keys, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # Yellow highlight on metadata columns for newly added runs
            if is_new and col_idx <= n_meta:
                cell.fill = NEW_META_FILL

            if key == "run_name":
                cell.value = row_data.get("run_name", "")
                url = row_data.get("_url", "")
                if url:
                    cell.hyperlink = url
                    cell.font = Font(color="FF0563C1", underline="single")

            elif key.startswith("_rank_"):
                prefix = key[len("_rank_"):]
                cell.value = ranks.get(prefix, {}).get(row_data.get("run_name"))
                cell.alignment = Alignment(horizontal="center")

            elif key.startswith("_delta_"):
                src_key = key[len("_delta_"):]  # e.g. "regression/rmse"
                v = row_data.get(src_key)
                b = baseline_row.get(src_key) if baseline_row else None
                if v is not None and b is not None:
                    cell.value = round(v - b, 6)

            else:
                val = row_data.get(key)
                cell.value = round(val, 6) if isinstance(val, float) else val

    last_data_row = len(rows) + 1  # last row index with actual data

    # ── Footer: MIN / MAX / MEAN ─────────────────────────────────────────────
    footer_start = last_data_row + 2  # one blank separator row
    for offset, label in enumerate(["MIN", "MAX", "MEAN"]):
        cell = ws.cell(row=footer_start + offset, column=1, value=label)
        cell.font = Font(bold=True)
        cell.fill = FOOTER_FILL

    for col_idx, key in enumerate(col_keys, start=1):
        # Apply footer formulas to metric columns (directional + delta)
        m_name = key.split("/")[-1]
        is_directional = (
            (not key.startswith("_") and _metric_direction(m_name) is not None)
            or key.startswith("_delta_")
        )
        if not is_directional:
            continue
        col_letter = get_column_letter(col_idx)
        data_range = f"{col_letter}2:{col_letter}{last_data_row}"
        for offset, fn in enumerate(["MIN", "MAX", "AVERAGE"]):
            cell = ws.cell(row=footer_start + offset, column=col_idx,
                           value=f"={fn}({data_range})")
            cell.fill = FOOTER_FILL

    # ── Conditional formatting (color scale per metric column) ───────────────
    for col_idx, key in enumerate(col_keys, start=1):
        if key.startswith("_delta_"):
            m_name = key.split("/")[-1]
            lower_b = m_name in LOWER_BETTER  # negative delta = improvement → green
        elif key.startswith("_"):
            continue  # _rank_ columns: no gradient
        else:
            lower_b = _metric_direction(key.split("/")[-1])
            if lower_b is None:
                continue  # neutral metric

        col_letter = get_column_letter(col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_data_row}",
            make_color_rule(lower_b),
        )

    # ── Bold the best value in each directional metric column ───────────────
    for col_idx, key in enumerate(col_keys, start=1):
        if key.startswith("_delta_"):
            m_name = key.split("/")[-1]
            lower_b = m_name in LOWER_BETTER
        elif key.startswith("_"):
            continue
        else:
            lower_b = _metric_direction(key.split("/")[-1])
            if lower_b is None:
                continue

        col_vals = [r.get(key) for r in rows]
        numeric = [v for v in col_vals if isinstance(v, (int, float))]
        if not numeric:
            continue
        best = min(numeric) if lower_b else max(numeric)
        for r_i, v in enumerate(col_vals):
            if v == best:
                ws.cell(row=r_i + 2, column=col_idx).font = Font(bold=True)
                break  # bold only the first occurrence

    # ── Freeze: top row + first 3 columns ───────────────────────────────────
    ws.freeze_panes = "D2"

    # ── Auto column widths (capped at 40) ────────────────────────────────────
    for col_idx, header in enumerate(col_headers, start=1):
        col_vals = [
            str(ws.cell(row=r, column=col_idx).value or "")
            for r in range(2, last_data_row + 1)
        ]
        max_len = max([len(header)] + [len(v) for v in col_vals])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # ── Hidden _meta sheet (stores timestamp + metrics for --update) ─────────
    ws_meta = wb.create_sheet("_meta")
    ws_meta.sheet_state = "hidden"
    ws_meta["A1"] = datetime.datetime.now(datetime.timezone.utc).isoformat()
    ws_meta["B1"] = "last_updated"
    ws_meta["A2"] = json.dumps(selected_metrics)
    ws_meta["B2"] = "selected_metrics"

    wb.save(output_path)
    print(f"  Saved {len(rows)} run(s) → {output_path}")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main():
    args = parse_args()

    # Resolve and validate selected metrics
    selected_metrics = [m.strip() for m in args.metrics.split(",") if m.strip()]
    unknown = [m for m in selected_metrics if m not in ALL_METRICS]
    if unknown:
        print(f"Warning: unknown metric(s) ignored: {unknown}")
        selected_metrics = [m for m in selected_metrics if m in ALL_METRICS]
    if not selected_metrics:
        sys.exit("Error: no valid metrics specified.")

    output_path = args.output or args.update or f"wandb_eval_runs_{datetime.date.today()}.xlsx"

    existing_rows = []
    existing_names: set = set()
    new_run_names:  set = set()
    user_set_metrics = (args.metrics != ",".join(DEFAULT_METRICS))

    # ── Update mode: load existing file ─────────────────────────────────────
    if args.update:
        if not os.path.exists(args.update):
            sys.exit(f"Error: --update file not found: {args.update}")
        print(f"Loading {args.update} ...")
        existing_rows, last_updated_str, saved_metrics = load_existing(args.update)
        existing_names = {r.get("run_name", "") for r in existing_rows}
        print(f"  {len(existing_rows)} existing run(s) loaded.")

        # Use saved metrics unless user explicitly passed --metrics
        if saved_metrics and not user_set_metrics:
            selected_metrics = saved_metrics
            print(f"  Using saved metrics from file: {selected_metrics}")

        # Auto-set --after from the last_updated timestamp stored in _meta
        if args.after is None and last_updated_str:
            try:
                dt = datetime.datetime.fromisoformat(str(last_updated_str))
                args.after = dt.date()
                print(f"  Fetching runs created after {args.after} (from _meta timestamp).")
            except (ValueError, TypeError):
                pass

    # ── Fetch new runs from WandB ────────────────────────────────────────────
    print("Fetching runs from WandB ...")
    runs = fetch_runs(
        entity=args.entity,
        project=args.project,
        state=args.state,
        after=args.after,
        before=args.before,
        max_runs=args.max_runs,
    )

    # Client-side tag filter (regex)
    if args.tag_filter:
        pat = re.compile(args.tag_filter)
        before_n = len(runs)
        runs = [r for r in runs if pat.search(r.config.get("run_tag", "") or "")]
        print(f"  Tag filter '{args.tag_filter}': {len(runs)}/{before_n} match.")

    # In update mode, skip runs already present in the file
    if args.update:
        new_only = [r for r in runs if r.name not in existing_names]
        new_run_names = {r.name for r in new_only}
        skipped = len(runs) - len(new_only)
        print(f"  {len(new_only)} new run(s) to add ({skipped} already present, skipped).")
        new_rows = [extract_row(r, selected_metrics) for r in new_only]
        all_rows = existing_rows + new_rows
    else:
        all_rows = [extract_row(r, selected_metrics) for r in runs]

    if not all_rows:
        print("No runs to export. Nothing written.")
        return

    # ── Write Excel ──────────────────────────────────────────────────────────
    print(f"Writing Excel ({len(all_rows)} run(s)) → {output_path} ...")
    write_excel(
        rows=all_rows,
        selected_metrics=selected_metrics,
        baseline_tag=args.baseline,
        output_path=output_path,
        new_run_names=new_run_names,
    )

    if new_run_names:
        print(f"  {len(new_run_names)} new run(s) highlighted in yellow (metadata columns).")


if __name__ == "__main__":
    main()
